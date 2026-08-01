import os
import time
import pandas as pd

from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import sys

START = int(sys.argv[1])
END = int(sys.argv[2])
output_dir = "data/output"
os.makedirs(output_dir, exist_ok=True)

excel_file = os.path.join(
    output_dir,
    f"medicine_details_{START}_{END}.xlsx"
)

if not os.path.exists(excel_file):

    wb = Workbook()
    ws = wb.active
    ws.title = "Medicines"

    ws.append([
        "Brand ID",

        "Medicine Name",
        "Medicine Name (Bangla)",

        "Dose Form",
        "Dose Form (Bangla)",

        "Generic Name",
        "Generic Name (Bangla)",

        "Strength",
        "Strength (Bangla)",

        "Manufacturer",
        "Manufacturer (Bangla)",

        "Unit Price",
        "Strip Price",

        "Indications",
        "Indications (Bangla)",

        "Pharmacology",
        "Pharmacology (Bangla)",

        "Dosage & Administration",
        "Dosage & Administration (Bangla)",

        "Interactions",
        "Interactions (Bangla)",

        "Contraindications",
        "Contraindications (Bangla)",

        "Side Effects",
        "Side Effects (Bangla)",

        "Pregnancy & Lactation",
        "Pregnancy & Lactation (Bangla)",

        "Precautions & Warnings",
        "Precautions & Warnings (Bangla)",

        "Use In Special Populations",
        "Use In Special Populations (Bangla)",

        "Therapeutic Class",
        "Therapeutic Class (Bangla)",

        "Storage Conditions",
        "Storage Conditions (Bangla)",

        "Image URL"
    ])

    wb.save(excel_file)

wb = load_workbook(excel_file)
ws = wb.active

csv_file = "data/medicine.csv"

read_csv = pd.read_csv(csv_file)
read_csv = read_csv.iloc[START:END]

options = Options()

options.add_argument("--start-maximized")
options.add_argument("--headless=new")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")
options.add_argument("--window-size=1920,1080")

options.set_capability(
    "goog:loggingPrefs",
    {
        "performance": "ALL",
        "browser": "ALL"
    }
)

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

wait = WebDriverWait(driver,10)

def safe_text(xpath):

    try:
        return driver.find_element(By.XPATH, xpath).text.strip()
    except:
        return ""


def safe_attr(xpath, attr):

    try:
        return driver.find_element(By.XPATH, xpath).get_attribute(attr)
    except:
        return ""


def safe_package():

    try:

        txt = driver.find_element(
            By.XPATH,
            "//*[contains(@class,'packages-wrapper')]"
        ).text

        arr = txt.split("\n")

        unit = arr[0] if len(arr) > 0 else ""
        strip = arr[1] if len(arr) > 1 else ""

        return unit, strip

    except:

        return "", ""
for index, row in read_csv.iterrows():

    brand_id = row["brand id"]

    print("=" * 80)
    print(f"Processing Brand ID : {brand_id}")

    try:

        url = f"https://medex.com.bd/brands/{brand_id}"

        driver.get(url)

        time.sleep(2)

        # এখানে Part-2 এর English Scraping Code থাকবে
        # ==========================
        # English Page
        # ==========================

        medicine_title = safe_text("//h1[@class='page-heading-1-l brand']")
        dose_form = safe_text("//small[@title='Dosage Form']")

        medicine_name = medicine_title.replace(dose_form, "").strip()

        generic_name = safe_text("//*[@title='Generic Name']")
        strength = safe_text("//*[@title='Strength']")
        manufacturer = safe_text("//*[@title='Manufactured by']")

        unitPrice, stripPrice = safe_package()

        indications = safe_text(
            "//*[@id='indications']//parent::*//*[@class='ac-body']"
        )

        pharmacology = safe_text(
            "//*[@id='mode_of_action']//parent::*//*[@class='ac-body']"
        )

        dosageAndAdministration = safe_text(
            "//*[@id='dosage']//parent::*//*[@class='ac-body']"
        )

        interactions = safe_text(
            "//*[@id='interaction']//parent::*//*[@class='ac-body']"
        )

        contraindications = safe_text(
            "//*[@id='contraindications']//parent::*//*[@class='ac-body']"
        )

        sideEffects = safe_text(
            "//*[@id='side_effects']//parent::*//*[@class='ac-body']"
        )

        pregnancyAndLactation = safe_text(
            "//*[@id='pregnancy_cat']//parent::*//*[@class='ac-body']"
        )

        precautionsAndWarnings = safe_text(
            "//*[@id='precautions']//parent::*//*[@class='ac-body']"
        )

        useInSpecialPopulations = safe_text(
            "//*[@id='pediatric_uses']//parent::*//*[@class='ac-body']"
        )

        therapeuticClass = safe_text(
            "//*[@id='drug_classes']//parent::*//*[@class='ac-body']"
        )

        storageConditions = safe_text(
            "//*[@id='storage_conditions']//parent::*//*[@class='ac-body']"
        )

        image_url = safe_attr(
            "//*[@title='Enlarge Pack Image']",
            "href"
        )
        # তারপর

        driver.get(driver.current_url + "/bn")

        time.sleep(2)

        # এখানে Part-2 এর Bangla Scraping Code থাকবে
        # ==========================
        # Bangla Page
        # ==========================


        medicine_title_bn = safe_text("//h1[@class='page-heading-1-l brand']")

        dose_form_bn = safe_text("//small[@title='Dosage Form']")

        medicine_name_bn = medicine_title_bn.replace(
            dose_form_bn,
            ""
        ).strip()

        generic_name_bn = safe_text("//*[@title='Generic Name']")

        strength_bn = safe_text("//*[@title='Strength']")

        manufacturer_bn = safe_text("//*[@title='Manufactured by']")

        indications_bn = safe_text(
            "//*[@id='indications']//parent::*//*[@class='ac-body']"
        )

        pharmacology_bn = safe_text(
            "//*[@id='mode_of_action']//parent::*//*[@class='ac-body']"
        )

        dosageAndAdministration_bn = safe_text(
            "//*[@id='dosage']//parent::*//*[@class='ac-body']"
        )

        interactions_bn = safe_text(
            "//*[@id='interaction']//parent::*//*[@class='ac-body']"
        )

        contraindications_bn = safe_text(
            "//*[@id='contraindications']//parent::*//*[@class='ac-body']"
        )

        sideEffects_bn = safe_text(
            "//*[@id='side_effects']//parent::*//*[@class='ac-body']"
        )

        pregnancyAndLactation_bn = safe_text(
            "//*[@id='pregnancy_cat']//parent::*//*[@class='ac-body']"
        )

        precautionsAndWarnings_bn = safe_text(
            "//*[@id='precautions']//parent::*//*[@class='ac-body']"
        )

        useInSpecialPopulations_bn = safe_text(
            "//*[@id='pediatric_uses']//parent::*//*[@class='ac-body']"
        )

        therapeuticClass_bn = safe_text(
            "//*[@id='drug_classes']//parent::*//*[@class='ac-body']"
        )

        storageConditions_bn = safe_text(
            "//*[@id='storage_conditions']//parent::*//*[@class='ac-body']"
        )
        time.sleep(2)
        ws.append([

            brand_id,

            medicine_name,
            medicine_name_bn,

            dose_form,
            dose_form_bn,

            generic_name,
            generic_name_bn,

            strength,
            strength_bn,

            manufacturer,
            manufacturer_bn,

            unitPrice,
            stripPrice,

            indications,
            indications_bn,

            pharmacology,
            pharmacology_bn,

            dosageAndAdministration,
            dosageAndAdministration_bn,

            interactions,
            interactions_bn,

            contraindications,
            contraindications_bn,

            sideEffects,
            sideEffects_bn,

            pregnancyAndLactation,
            pregnancyAndLactation_bn,

            precautionsAndWarnings,
            precautionsAndWarnings_bn,

            useInSpecialPopulations,
            useInSpecialPopulations_bn,

            therapeuticClass,
            therapeuticClass_bn,

            storageConditions,
            storageConditions_bn,

            image_url

        ])
        wb.save(excel_file)

        print("Saved.")
    except Exception as e:

        print(f"Failed : {brand_id}")

        print(e)

        continue
    wb.save(excel_file)

    wb.close()

driver.quit()

print("=" * 80)
print("Completed Successfully.")
